#!/usr/bin/env python
# coding: utf-8
"""
confusion_matrix.py
-------------------
Confusion matrix generation for grade_predictor models.

Usage (standalone):
    python confusion_matrix.py

This script loads the saved model checkpoints from result/classification/
and result/ordinal/, evaluates each model on the validation split
(random_state=42 keeps it identical to training), and saves mean
row-normalized confusion matrix images + an Excel sheet.

Current default behavior evaluates split-based checkpoints under
result/split_models across all splits listed in result/split_indices.json.
You can filter by --split-id and --model-name.

Usage (as a library):
    from confusion_matrix import plot_mean_confusion_matrix
    plot_mean_confusion_matrix(cm_records)
"""

import glob
import os
import json
import argparse

import matplotlib.pyplot as plt
import numpy as np
import seaborn as sns
import torch
from collections import defaultdict
from sklearn.metrics import confusion_matrix
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# ---------------------------------------------------------------------------
# Core plotting / saving helpers
# ---------------------------------------------------------------------------

def plot_mean_confusion_matrix(
    all_cm_records,
    class_labels=None,
    class_indices=None,
    excel_path="result/model_confusion_matrices.xlsx",
    img_dir="result",
    embed_figure_sheets=False,
):
    """
    For each model, average the row-normalized confusion matrices across all
    records, save a PNG and write each matrix into its own sheet in one Excel.

    Parameters
    ----------
    all_cm_records : list[dict]
        Each dict must contain:
            "model"  : str              – model name
            "y_true" : list/array[int]  – true class indices
            "y_pred" : list/array[int]  – predicted class indices
    class_labels : list[str], optional
        Display labels for each class.
    class_indices : list[int], optional
        Class indices used to compute confusion matrix. Must align with class_labels.
    excel_path : str
        Excel file path to write confusion matrices.
    img_dir : str
        Directory for PNG output.
    embed_figure_sheets : bool
        If True, also add image tabs to the Excel workbook.
    """
    if not all_cm_records:
        print("No confusion matrix records supplied; skipping.")
        return
    if class_labels is None:
        class_labels = [f"V{i}" for i in range(4, 11)]
    if class_indices is None:
        class_indices = list(range(len(class_labels)))
    if len(class_labels) != len(class_indices):
        raise ValueError("class_labels and class_indices must have the same length")

    os.makedirs(img_dir, exist_ok=True)

    cm_by_model = defaultdict(list)
    for record in all_cm_records:
        y_true = np.asarray(record["y_true"])
        y_pred = np.asarray(record["y_pred"])
        cm_by_model[record["model"]].append((y_true, y_pred))

    image_paths = {}

    for model_name, runs in cm_by_model.items():
        cms = []
        for y_true, y_pred in runs:
            cm = confusion_matrix(
                y_true, y_pred, labels=class_indices, normalize="true"
            )
            cms.append(cm)
        mean_cm = np.mean(cms, axis=0)

        fig, ax = plt.subplots(figsize=(8, 6))
        sns.heatmap(
            mean_cm,
            annot=True,
            fmt=".2f",
            cmap="Blues",
            xticklabels=class_labels,
            yticklabels=class_labels,
            ax=ax,
        )
        n_runs = len(runs)
        ax.set_title(f"Mean Confusion Matrix ({n_runs} run{'s' if n_runs != 1 else ''}): {model_name}")
        ax.set_xlabel("Predicted Grade")
        ax.set_ylabel("Actual Grade")
        plt.tight_layout()

        safe_name = "".join(
            ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in model_name
        )
        img_path = os.path.join(img_dir, f"confusion_mean_{safe_name}.png")
        plt.savefig(img_path, dpi=150)
        plt.close(fig)
        image_paths[model_name] = img_path
        print(f"Saved confusion matrix image: {img_path}")

    # Write one workbook with one sheet per model (tab = model).
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        used_sheet_names = set()
        for model_name, runs in cm_by_model.items():
            cms = []
            for y_true, y_pred in runs:
                cm = confusion_matrix(
                    y_true, y_pred, labels=class_indices, normalize="true"
                )
                cms.append(cm)
            mean_cm = np.mean(cms, axis=0)

            df_cm = pd.DataFrame(mean_cm, index=class_labels, columns=class_labels)
            safe_name = "".join(
                ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in model_name
            )
            base_sheet_name = safe_name[:31] if safe_name else "model"
            if not base_sheet_name:
                base_sheet_name = "model"

            sheet_name = base_sheet_name
            suffix = 2
            while sheet_name in used_sheet_names:
                tail = f"_{suffix}"
                sheet_name = (base_sheet_name[: 31 - len(tail)] + tail)[:31]
                suffix += 1
            used_sheet_names.add(sheet_name)

            df_cm.to_excel(writer, sheet_name=sheet_name, startrow=2)
            ws = writer.book[sheet_name]
            ws["A1"] = f"Model: {model_name}"
            ws["A2"] = f"Num runs (splits): {len(runs)}"

    print(f"Saved model-wise confusion matrix workbook: {excel_path}")

    if embed_figure_sheets and image_paths:
        wb = load_workbook(excel_path)
        used_sheet_names = set(wb.sheetnames)
        for model_name, img_path in image_paths.items():
            safe_name = "".join(
                ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in model_name
            )
            base_sheet_name = (safe_name + "_fig")[:31] if safe_name else "model_fig"
            if not base_sheet_name:
                base_sheet_name = "model_fig"

            sheet_name = base_sheet_name
            suffix = 2
            while sheet_name in used_sheet_names:
                tail = f"_{suffix}"
                sheet_name = (base_sheet_name[: 31 - len(tail)] + tail)[:31]
                suffix += 1
            used_sheet_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            ws["A1"] = f"Model: {model_name}"
            ws["A2"] = "Confusion matrix figure"
            ws.add_image(XLImage(img_path), "A4")

        wb.save(excel_path)
        print("Saved figure tabs in Excel workbook.")


# ---------------------------------------------------------------------------
# Standalone evaluation helpers
# ---------------------------------------------------------------------------

def _get_device():
    if torch.cuda.is_available():
        return torch.device("cuda")
    if torch.backends.mps.is_available():
        return torch.device("mps")
    return torch.device("cpu")


def _predict(model, val_loader, device):
    """Return (y_true, y_pred) lists for a val_loader."""
    from models.utils_ordinal import cumulative_to_labels

    model.eval()
    y_true_all, y_pred_all = [], []
    with torch.no_grad():
        for X, y in val_loader:
            inputs = tuple(x.to(device) for x in X)
            y = y.to(device)
            payload = inputs[0] if len(inputs) == 1 else inputs
            outputs = model(payload)

            if isinstance(outputs, tuple):
                if getattr(model, "is_ordinal", False):
                    probs = outputs[0]
                    preds = cumulative_to_labels(probs)
                else:
                    probs = outputs[0]
                    preds = probs.argmax(dim=1)
            else:
                preds = outputs.argmax(dim=1)

            y_true_all.extend(y.cpu().numpy())
            y_pred_all.extend(preds.detach().cpu().numpy())

    return y_true_all, y_pred_all


def _build_val_loader(model_type, dataset, val_idx, batch_size=16):
    """Create a DataLoader for the validation subset."""
    from torch.utils.data import DataLoader, Subset
    from main import make_collate_fn  # thin import – no side effects beyond globals

    collate_fn = make_collate_fn(model_type)
    val_subset = Subset(dataset, val_idx)
    return DataLoader(
        val_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn
    )


def _load_splits(splits_path):
    with open(splits_path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    splits = payload.get("splits", [])
    if not splits:
        raise ValueError(f"No splits found in {splits_path}")
    return splits


def _resolve_splits_path(splits_path):
    candidates = [splits_path, "result/split_indices.json", "result/splits/split_indices.json"]
    for path in candidates:
        if path and os.path.exists(path):
            return path
    raise FileNotFoundError(
        "Split index JSON was not found. Checked: " + ", ".join(candidates)
    )


def _infer_model_type(model_name, model_obj):
    """Infer collate model type from model name + model object characteristics."""
    lower_name = model_name.lower()
    class_name = model_obj.__class__.__name__.lower()

    is_ordinal = (
        ("ordinal" in lower_name)
        or ("ordinal" in class_name)
        or bool(getattr(model_obj, "is_ordinal", False))
    )
    uses_xy = (
        ("_xy" in lower_name)
        or ("additive" in lower_name)
        or ("xy" in class_name)
        or bool(getattr(model_obj, "expects_tuple_input", False))
    )

    members = getattr(model_obj, "models", None)
    if members is not None:
        try:
            for member in members.values():
                if bool(getattr(member, "expects_tuple_input", False)):
                    uses_xy = True
                    break
        except Exception:
            pass

    if is_ordinal and uses_xy:
        return "set_transformer_ordinal_xy"
    if is_ordinal:
        return "set_transformer_ordinal"
    if uses_xy:
        return "set_transformer_xy"
    return "set_transformer"


def load_and_evaluate_checkpoints(
    model_root="result/split_models",
    splits_path="result/split_indices.json",
    split_id=None,
    model_name=None,
    json_path="./data/cleaned_moonboard2024_grouped.json",
    excel_path="result/model_confusion_matrices.xlsx",
    img_dir="result",
    batch_size=16,
):
    """
    Evaluate split-based checkpoints under model_root and plot model-wise
    mean confusion matrices. Defaults to all splits and all models.
    """
    # ---- Import dataset utilities from main (runs module-level setup once) --
    import main as M  # noqa: E402  (intentional late import)

    device = _get_device()
    print(f"Using device: {device}")

    # ---- Build dataset once -------------------------------------------------
    dataset = M.load_dataset(
        json_path,
        M.hold_to_idx,
        M.grade_to_label,
        M.hold_difficulty,
        M.type_to_idx,
        M.hold_to_coord,
    )

    present_class_indices = sorted(
        {int(M.grade_to_label[item["grade"]]) for item in dataset.raw}
    )
    class_labels = [
        M.label_to_grade.get(idx, f"Class_{idx}") for idx in present_class_indices
    ]
    print(f"Using class labels: {class_labels}")

    splits_path = _resolve_splits_path(splits_path)
    splits = _load_splits(splits_path)

    if split_id is not None:
        splits = [s for s in splits if int(s.get("split_id", -1)) == int(split_id)]
        if not splits:
            raise ValueError(f"split_id={split_id} was not found in {splits_path}")

    # ---- Collect predictions for each checkpoint ----------------------------
    cm_records = []
    for split in splits:
        cur_split_id = int(split.get("split_id"))
        val_idx = [int(x) for x in split.get("test_idx", [])]
        split_dir = os.path.join(model_root, f"split_{cur_split_id:03d}")

        if not os.path.isdir(split_dir):
            print(f"Directory not found, skipping split {cur_split_id}: {split_dir}")
            continue

        pt_files = []
        for subdir in ("classification", "ordinal"):
            ckpt_dir = os.path.join(split_dir, subdir)
            if os.path.isdir(ckpt_dir):
                pt_files.extend(sorted(glob.glob(os.path.join(ckpt_dir, "*.pt"))))

        if not pt_files:
            print(f"No .pt files found in split {cur_split_id}: {split_dir}")
            continue

        for pt_path in pt_files:
            current_model_name = os.path.splitext(os.path.basename(pt_path))[0]
            if model_name and current_model_name != model_name:
                continue

            print(f"Loading checkpoint: {pt_path}")
            try:
                try:
                    payload = torch.load(pt_path, map_location=device, weights_only=False)
                except TypeError:
                    payload = torch.load(pt_path, map_location=device)
                model = payload.get("model_object")
                if model is None:
                    print(f"  No 'model_object' key; skipping {current_model_name}")
                    continue
                model = model.to(device)
                model.eval()
            except Exception as exc:
                print(f"  Failed to load {pt_path}: {exc}")
                continue

            collate_type = _infer_model_type(current_model_name, model)
            try:
                val_loader = _build_val_loader(
                    collate_type, dataset, val_idx, batch_size=batch_size
                )
                y_true, y_pred = _predict(model, val_loader, device)
            except Exception as exc:
                print(f"  Evaluation failed for {current_model_name}: {exc}")
                continue

            cm_records.append(
                {
                    "model": current_model_name,
                    "split_id": cur_split_id,
                    "y_true": y_true,
                    "y_pred": y_pred,
                }
            )
            print(
                f"  Collected predictions for {current_model_name} "
                f"(split={cur_split_id}, {len(y_true)} samples)"
            )

    # ---- Plot ---------------------------------------------------------------
    if cm_records:
        plot_mean_confusion_matrix(
            cm_records,
            class_labels=class_labels,
            class_indices=present_class_indices,
            excel_path=excel_path,
            img_dir=img_dir,
            embed_figure_sheets=bool(model_name),
        )
    else:
        print("No valid predictions collected; no confusion matrices generated.")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=(
            "Generate confusion matrices for split-based checkpoints. "
            "Default: evaluate all models in all splits under result/split_models."
        )
    )
    parser.add_argument("--model-root", type=str, default="result/split_models")
    parser.add_argument("--splits-path", type=str, default="result/split_indices.json")
    parser.add_argument(
        "--split-id",
        type=int,
        default=None,
        help="Target split id (e.g., 1). If omitted, all splits are used.",
    )
    parser.add_argument(
        "--model-name",
        type=str,
        default=None,
        help="Target model name (e.g., xgboost_ensemble_all). If omitted, all models are used.",
    )
    parser.add_argument(
        "--json-path",
        type=str,
        default="./data/cleaned_moonboard2024_grouped.json",
    )
    parser.add_argument(
        "--excel-path",
        type=str,
        default="result/model_confusion_matrices.xlsx",
    )
    parser.add_argument("--img-dir", type=str, default="result")
    parser.add_argument("--batch-size", type=int, default=16)
    args = parser.parse_args()

    load_and_evaluate_checkpoints(
        model_root=args.model_root,
        splits_path=args.splits_path,
        split_id=args.split_id,
        model_name=args.model_name,
        json_path=args.json_path,
        excel_path=args.excel_path,
        img_dir=args.img_dir,
        batch_size=args.batch_size,
    )
